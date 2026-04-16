import streamlit as st
st.set_page_config(page_title="Template CM | INTERLOG", page_icon="📋", layout="wide")
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
import io, math, zipfile

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap');
    html, body, [class*="css"] { font-family: 'Roboto', sans-serif; background-color: #0e1117; color: #fafafa; }
    .block-container { padding-top: 3rem; }
    h1 { color: #00b4d8 !important; font-size: 1.8rem !important; }
    .subtitulo-app { font-size: 0.95rem; color: #8899aa; margin-bottom: 1.5rem; }
    .seccion-titulo { font-size: 1rem; font-weight: 600; color: #00b4d8; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 1rem; border-bottom: 1px solid #1e3a4a; padding-bottom: 6px; }
    .badge-ok { background: #0d2e1a; color: #3dd68c; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #3dd68c; display: inline-block; margin: 3px 2px; }
    .badge-err { background: #2e0d0d; color: #ff6b6b; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #ff6b6b; display: inline-block; margin: 3px 2px; }
    .badge-fixed { background: #1a2a1a; color: #90ee90; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #90ee90; display: inline-block; margin: 3px 2px; }
    .excluido-box { background: #1a1000; border: 1px solid #f0c040; border-radius: 8px; padding: 12px 16px; margin: 4px 0; font-size: 0.85rem; color: #ffd980; }
    .excluido-title { color: #f0c040; font-weight: 600; margin-bottom: 8px; }
    .no-genera-box { background: #2e0d0d; border: 1px solid #ff6b6b; border-radius: 8px; padding: 12px 16px; margin: 4px 0; font-size: 0.85rem; color: #ff6b6b; }
    .alerta-descarte { background: #2a1f00; border: 1px solid #f0c040; border-radius: 8px; padding: 12px 16px; color: #f0c040; font-size: 0.88rem; margin: 6px 0; }
    .stButton > button { background: linear-gradient(135deg, #00b4d8, #0077b6) !important; color: white !important; font-weight: 600 !important; border: none !important; border-radius: 8px !important; }
    div[data-testid="stFileUploader"] label p { color: #00b4d8 !important; font-weight: 500 !important; }
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

ARCHIVOS_FIJOS = {'razon': 'RAZON_SOCIAL_FSM.xlsx', 'proyectos': 'PROYECTOS.xlsx', 'ncm': 'NCM_LEY_MINERA.xlsx'}

TRADUCCION_ORIGEN = {
    "USA": "ESTADOS UNIDOS", "UNITED STATES": "ESTADOS UNIDOS", "UNITED STATES OF AMERICA": "ESTADOS UNIDOS",
    "MEXICO": "MÉXICO", "CANADA": "CANADÁ", "GUATEMALA": "GUATEMALA", "BELIZE": "BELICE",
    "HONDURAS": "HONDURAS", "EL SALVADOR": "EL SALVADOR", "NICARAGUA": "NICARAGUA",
    "COSTA RICA": "COSTA RICA", "PANAMA": "PANAMÁ", "CUBA": "CUBA", "HAITI": "HAITÍ",
    "DOMINICAN REPUBLIC": "REPÚBLICA DOMINICANA", "PUERTO RICO": "PUERTO RICO", "JAMAICA": "JAMAICA",
    "TRINIDAD AND TOBAGO": "TRINIDAD Y TOBAGO", "BARBADOS": "BARBADOS", "BRAZIL": "BRASIL",
    "ARGENTINA": "ARGENTINA", "CHILE": "CHILE", "PERU": "PERÚ", "COLOMBIA": "COLOMBIA",
    "VENEZUELA": "VENEZUELA", "ECUADOR": "ECUADOR", "BOLIVIA": "BOLIVIA", "PARAGUAY": "PARAGUAY",
    "URUGUAY": "URUGUAY", "GUYANA": "GUYANA", "SURINAME": "SURINAM", "FRENCH GUIANA": "GUAYANA FRANCESA",
    "GERMANY": "ALEMANIA", "ITALY": "ITALIA", "FRANCE": "FRANCIA", "SPAIN": "ESPAÑA",
    "UNITED KINGDOM": "REINO UNIDO", "UK": "REINO UNIDO", "GREAT BRITAIN": "REINO UNIDO",
    "ENGLAND": "REINO UNIDO", "BELGIUM": "BÉLGICA", "NETHERLANDS": "PAÍSES BAJOS",
    "HOLLAND": "PAÍSES BAJOS", "LUXEMBOURG": "LUXEMBURGO", "SWITZERLAND": "SUIZA",
    "AUSTRIA": "AUSTRIA", "PORTUGAL": "PORTUGAL", "IRELAND": "IRLANDA", "SWEDEN": "SUECIA",
    "NORWAY": "NORUEGA", "DENMARK": "DINAMARCA", "FINLAND": "FINLANDIA", "ICELAND": "ISLANDIA",
    "MONACO": "MÓNACO", "LIECHTENSTEIN": "LIECHTENSTEIN", "ANDORRA": "ANDORRA", "MALTA": "MALTA",
    "CYPRUS": "CHIPRE", "POLAND": "POLONIA", "CZECH REPUBLIC": "REPÚBLICA CHECA",
    "CZECHIA": "REPÚBLICA CHECA", "SLOVAKIA": "ESLOVAQUIA", "HUNGARY": "HUNGRÍA",
    "ROMANIA": "RUMANIA", "BULGARIA": "BULGARIA", "CROATIA": "CROACIA", "SLOVENIA": "ESLOVENIA",
    "SERBIA": "SERBIA", "BOSNIA AND HERZEGOVINA": "BOSNIA Y HERZEGOVINA", "MONTENEGRO": "MONTENEGRO",
    "NORTH MACEDONIA": "MACEDONIA DEL NORTE", "ALBANIA": "ALBANIA", "GREECE": "GRECIA",
    "ESTONIA": "ESTONIA", "LATVIA": "LETONIA", "LITHUANIA": "LITUANIA", "UKRAINE": "UCRANIA",
    "BELARUS": "BIELORRUSIA", "MOLDOVA": "MOLDAVIA", "RUSSIA": "RUSIA",
    "RUSSIAN FEDERATION": "RUSIA", "GEORGIA": "GEORGIA", "ARMENIA": "ARMENIA",
    "AZERBAIJAN": "AZERBAIYÁN", "KOSOVO": "KOSOVO", "CHINA": "CHINA", "JAPAN": "JAPÓN",
    "SOUTH KOREA": "COREA DEL SUR", "KOREA": "COREA DEL SUR", "REPUBLIC OF KOREA": "COREA DEL SUR",
    "NORTH KOREA": "COREA DEL NORTE", "TAIWAN": "TAIWÁN", "HONG KONG": "HONG KONG",
    "MACAO": "MACAO", "MONGOLIA": "MONGOLIA", "VIETNAM": "VIETNAM", "THAILAND": "TAILANDIA",
    "MALAYSIA": "MALASIA", "SINGAPORE": "SINGAPUR", "INDONESIA": "INDONESIA",
    "PHILIPPINES": "FILIPINAS", "MYANMAR": "MYANMAR", "CAMBODIA": "CAMBOYA", "LAOS": "LAOS",
    "BRUNEI": "BRUNÉI", "TIMOR-LESTE": "TIMOR ORIENTAL", "INDIA": "INDIA", "PAKISTAN": "PAKISTÁN",
    "BANGLADESH": "BANGLADÉS", "SRI LANKA": "SRI LANKA", "NEPAL": "NEPAL", "BHUTAN": "BUTÁN",
    "MALDIVES": "MALDIVAS", "AFGHANISTAN": "AFGANISTÁN", "KAZAKHSTAN": "KAZAJISTÁN",
    "UZBEKISTAN": "UZBEKISTÁN", "TURKMENISTAN": "TURKMENISTÁN", "KYRGYZSTAN": "KIRGUISTÁN",
    "TAJIKISTAN": "TAYIKISTÁN", "TURKEY": "TURQUÍA", "TURKIYE": "TURQUÍA", "ISRAEL": "ISRAEL",
    "SAUDI ARABIA": "ARABIA SAUDITA", "UNITED ARAB EMIRATES": "EMIRATOS ÁRABES UNIDOS",
    "UAE": "EMIRATOS ÁRABES UNIDOS", "QATAR": "CATAR", "KUWAIT": "KUWAIT", "BAHRAIN": "BARÉIN",
    "OMAN": "OMÁN", "JORDAN": "JORDANIA", "LEBANON": "LÍBANO", "SYRIA": "SIRIA",
    "IRAQ": "IRAK", "IRAN": "IRÁN", "YEMEN": "YEMEN", "EGYPT": "EGIPTO", "LIBYA": "LIBIA",
    "TUNISIA": "TÚNEZ", "ALGERIA": "ARGELIA", "MOROCCO": "MARRUECOS", "SUDAN": "SUDÁN",
    "SOUTH AFRICA": "SUDÁFRICA", "NIGERIA": "NIGERIA", "KENYA": "KENYA", "ETHIOPIA": "ETIOPÍA",
    "GHANA": "GHANA", "TANZANIA": "TANZANIA", "MOZAMBIQUE": "MOZAMBIQUE", "ZAMBIA": "ZAMBIA",
    "ZIMBABWE": "ZIMBABUE", "ANGOLA": "ANGOLA",
    "DEMOCRATIC REPUBLIC OF THE CONGO": "REPÚBLICA DEMOCRÁTICA DEL CONGO",
    "DRC": "REPÚBLICA DEMOCRÁTICA DEL CONGO", "CONGO": "CONGO", "CAMEROON": "CAMERÚN",
    "IVORY COAST": "COSTA DE MARFIL", "SENEGAL": "SENEGAL", "MALI": "MALÍ",
    "BURKINA FASO": "BURKINA FASO", "NIGER": "NÍGER", "CHAD": "CHAD", "SOMALIA": "SOMALIA",
    "UGANDA": "UGANDA", "RWANDA": "RUANDA", "BOTSWANA": "BOTSUANA", "NAMIBIA": "NAMIBIA",
    "MADAGASCAR": "MADAGASCAR", "AUSTRALIA": "AUSTRALIA", "NEW ZEALAND": "NUEVA ZELANDA",
    "PAPUA NEW GUINEA": "PAPÚA NUEVA GUINEA", "FIJI": "FIYI",
    "SOLOMON ISLANDS": "ISLAS SALOMÓN", "VANUATU": "VANUATU", "SAMOA": "SAMOA",
    "TONGA": "TONGA", "KIRIBATI": "KIRIBATI", "NEW CALEDONIA": "NUEVA CALEDONIA",
}

UNIDAD_MAP = {'07 - UNIDAD': 'UNIDAD', '01 - KILOGRAMO': 'KILOGRAMO', '06 - METRO': 'METRO', '10 - LITRO': 'LITRO'}
COLUMNAS = [
    'ID','InscRUMP','ActiServ','NroInsc','RazonSocial','CUIT','ImpDirecta','CondMerca',
    'AnioFabricacion','RazonSocialLeasing','CuitLeasing','SimiSira','ValorFOBTotal',
    'ProyectoMinero','Radicacion','DetalleTransitorioDeposito','ClasificacionDeArticulo',
    'TipoDeFactura','NumeroDeFactura','OrdenDeCompra','Descripcion','Cantidad',
    'UnidadMedida','PosicionArancelaria','Marca','Modelo','NroDeSerie','ValorUnitario',
    'ValorTotalItem','SerieDelMotor','MarcaDelMotor','ModeloMotor','TipoPlantaDestino',
    'FinalidadDeUso','CodigoParte','TipoDeMaquina','MarcaMaquina','ModeloMaquina',
    'Expedientes Escalonados','Proveedor','PaisOrigenDeLaMercaderia','Observaciones','ITEM_DESPACHO'
]
SIN_COLOR = {
    'ID','InscRUMP','ActiServ','NroInsc','RazonSocial','CUIT','ImpDirecta','CondMerca',
    'SimiSira','ProyectoMinero','Radicacion','ClasificacionDeArticulo','TipoDeFactura',
    'Observaciones','ITEM_DESPACHO'
}

def safe_float(v):
    try:
        f = float(v); return 0.0 if math.isnan(f) else f
    except: return 0.0

def find_col(df, *kw_groups):
    for kw_group in kw_groups:
        if isinstance(kw_group, str):
            kw_group = [kw_group]
        for col in df.columns:
            col_up = col.upper().strip().replace(' ', '').replace('_', '')
            if all(kw.upper().replace(' ', '').replace('_', '') in col_up for kw in kw_group):
                return col
    return None

def norm_pn(pn):
    return str(pn).strip().lstrip('0')

def traducir_pais(raw, traducciones_extra=None):
    r = str(raw).strip().upper()
    if traducciones_extra and r in traducciones_extra:
        return traducciones_extra[r]
    return TRADUCCION_ORIGEN.get(r, r)

def parsear_equipo(eq):
    if not eq or str(eq).strip() in ('', 'nan'): return '', 'CAT', ''
    eq = str(eq).strip()
    if ' - ' in eq:
        p = eq.split(' - ', 1); return p[0].strip(), 'CAT', p[1].strip()
    tokens = eq.split()
    for i, t in enumerate(tokens):
        if t.upper() in ['CAT', 'CATERPILLAR']:
            return ' '.join(tokens[:i]).strip(), 'CAT', ' '.join(tokens[i+1:]).strip()
    return (' '.join(tokens[:-1]).strip(), 'CAT', tokens[-1].strip()) if len(tokens) > 1 else (eq, 'CAT', '')

def calcular_fobs(items, col_ext='EXTENDED_PRICE', col_spk='SPECIAL PACKING',
                  col_frt='FREIGHT_CHARGE', col_bofrt='BO_FREIGHT_CHARGE',
                  col_emerg='EMERGENCY_FILL_CHARGE_VAL'):
    total_ext = sum(safe_float(i.get(col_ext)) for i in items)
    total_gastos = sum(
        safe_float(i.get(col_spk)) + safe_float(i.get(col_frt)) +
        safe_float(i.get(col_bofrt)) + safe_float(i.get(col_emerg))
        for i in items
    )
    return [round(safe_float(i.get(col_ext)) + (safe_float(i.get(col_ext)) * total_gastos / total_ext if total_ext else 0), 2) for i in items]

def match_subitem(pn, qty, fob_calc, df_sub, usados):
    pn_n = norm_pn(pn)
    try: qty_val = float(qty)
    except: return None, None, 'sin_match'
    mask = (df_sub['MODELO_NORM'] == pn_n) & (df_sub['CANTIDAD'].astype(float) == qty_val)
    disp = df_sub[mask & ~df_sub.index.isin(usados)].copy()
    if disp.empty: return None, None, 'sin_match'
    m = disp[(disp['MONTO FOB'].astype(float) - fob_calc).abs() <= 0.02]
    if not m.empty:
        idx = m.index[0]; usados.add(idx); return m.loc[idx], str(m.loc[idx]['ITEM']), 'unico'
    idx = disp.index[0]; usados.add(idx); return disp.loc[idx], str(disp.loc[idx]['ITEM']), 'descarte'

def generar_excel_bytes(filas):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'general'
    FILL_AM = PatternFill('solid', fgColor='FFFF00')
    FH = Font(name='Arial', size=11, bold=True)
    FD = Font(name='Calibri', size=11)
    for ci, cn in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = FH; cell.alignment = Alignment(horizontal='left', vertical='center')
        if cn not in SIN_COLOR: cell.fill = FILL_AM
    for ri, fila in enumerate(filas, 2):
        for ci, cn in enumerate(COLUMNAS, 1):
            cell = ws.cell(row=ri, column=ci, value=fila.get(cn))
            cell.font = FD; cell.alignment = Alignment(vertical='center')
    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 10), 45)
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def construir_filas_excluidos(excluidos, inv, rs, cuit_val, cond_merca):
    filas = []
    for f in excluidos:
        p = f['proy']
        filas.append({
            'ID': 0, 'InscRUMP': str(rs['InscRUMP']), 'ActiServ': str(rs['ActiServ']),
            'NroInsc': str(rs['NroInsc']), 'RazonSocial': str(rs['RazonSocial']),
            'CUIT': cuit_val, 'ImpDirecta': str(rs['ImpDirecta']), 'CondMerca': cond_merca,
            'AnioFabricacion': None, 'RazonSocialLeasing': None, 'CuitLeasing': None, 'SimiSira': None,
            'ValorFOBTotal': f['fob_final'],
            'ProyectoMinero': str(p.get('ProyectoMinero', '')).upper() if p else '',
            'Radicacion': str(p.get('Radicacion', '')).upper() if p else '',
            'DetalleTransitorioDeposito': str(p.get('DetalleTransitorioDeposito', '')) if p else '',
            'ClasificacionDeArticulo': str(rs['ClasificacionDeArticulo']),
            'TipoDeFactura': 'DEFINITIVA', 'NumeroDeFactura': inv, 'OrdenDeCompra': None,
            'Descripcion': f['descripcion'], 'Cantidad': f['qty_int'], 'UnidadMedida': f['unidad'],
            'PosicionArancelaria': f['ncm'], 'Marca': None, 'Modelo': None,
            'NroDeSerie': None, 'ValorUnitario': f['valor_unit'], 'ValorTotalItem': f['valor_total'],
            'SerieDelMotor': None, 'MarcaDelMotor': None, 'ModeloMotor': None,
            'TipoPlantaDestino': None, 'FinalidadDeUso': None,
            'CodigoParte': f['pn'], 'TipoDeMaquina': f['tipo_maq'],
            'MarcaMaquina': f['marca_maq'], 'ModeloMaquina': f['modelo_maq'],
            'Expedientes Escalonados': None, 'Proveedor': None,
            'PaisOrigenDeLaMercaderia': f['origen'], 'Observaciones': 'SIN CM',
            'ITEM_DESPACHO': f['item_di'],
        })
    return filas

def construir_filas(grupo_items, inv, rs, cuit_val, cond_merca):
    valor_fob_total = round(sum(i['fob_final'] for i in grupo_items if i['fob_final']), 2)
    filas = []
    for f in grupo_items:
        p = f['proy']
        filas.append({
            'ID': 0, 'InscRUMP': str(rs['InscRUMP']), 'ActiServ': str(rs['ActiServ']),
            'NroInsc': str(rs['NroInsc']), 'RazonSocial': str(rs['RazonSocial']),
            'CUIT': cuit_val, 'ImpDirecta': str(rs['ImpDirecta']), 'CondMerca': cond_merca,
            'AnioFabricacion': None, 'RazonSocialLeasing': None, 'CuitLeasing': None, 'SimiSira': None,
            'ValorFOBTotal': valor_fob_total,
            'ProyectoMinero': str(p.get('ProyectoMinero', '')).upper() if p else '',
            'Radicacion': str(p.get('Radicacion', '')).upper() if p else '',
            'DetalleTransitorioDeposito': str(p.get('DetalleTransitorioDeposito', '')) if p else '',
            'ClasificacionDeArticulo': str(rs['ClasificacionDeArticulo']),
            'TipoDeFactura': 'DEFINITIVA', 'NumeroDeFactura': inv, 'OrdenDeCompra': None,
            'Descripcion': f['descripcion'], 'Cantidad': f['qty_int'], 'UnidadMedida': f['unidad'],
            'PosicionArancelaria': f['ncm'], 'Marca': 'CATERPILLAR', 'Modelo': 'NO POSEE',
            'NroDeSerie': 'NO POSEE', 'ValorUnitario': f['valor_unit'], 'ValorTotalItem': f['valor_total'],
            'SerieDelMotor': None, 'MarcaDelMotor': None, 'ModeloMotor': None,
            'TipoPlantaDestino': None, 'FinalidadDeUso': None,
            'CodigoParte': f['pn'], 'TipoDeMaquina': f['tipo_maq'],
            'MarcaMaquina': f['marca_maq'], 'ModeloMaquina': f['modelo_maq'],
            'Expedientes Escalonados': 'no', 'Proveedor': 'CATERPILLAR',
            'PaisOrigenDeLaMercaderia': f['origen'], 'Observaciones': None,
            'ITEM_DESPACHO': f['item_di'],
        })
    return filas

def procesar(f_madre, f_despacho, f_equipos, f_desc, cond_merca):
    xl_madre = pd.ExcelFile(f_madre)
    sheet_madre = 'Hoja1' if 'Hoja1' in xl_madre.sheet_names else xl_madre.sheet_names[0]
    df_madre = pd.read_excel(f_madre, sheet_name=sheet_madre, dtype=str)
    df_sub = pd.read_excel(f_despacho, sheet_name='Subitem', dtype=str)
    df_liq = pd.read_excel(f_despacho, sheet_name='Liquidación ítem', dtype=str)
    xl_eq = pd.ExcelFile(f_equipos)
    sheet_eq = xl_eq.sheet_names[0]
    for sh in xl_eq.sheet_names:
        try:
            cols_sh = [str(c).strip() for c in pd.read_excel(f_equipos, sheet_name=sh, nrows=1, dtype=str).columns]
            cols_up = [c.upper() for c in cols_sh]
            if any('PART' in c for c in cols_up) and any('EQUIP' in c for c in cols_up):
                sheet_eq = sh; break
            if sh.lower() == 'data':
                sheet_eq = sh; break
        except: pass
    df_eq = pd.read_excel(f_equipos, sheet_name=sheet_eq, dtype=str)
    xl_desc = pd.ExcelFile(f_desc)
    sheet_desc = xl_desc.sheet_names[0]
    for sh in xl_desc.sheet_names:
        try:
            cols = [str(c).strip().upper() for c in pd.read_excel(f_desc, sheet_name=sh, nrows=1, dtype=str).columns]
            if any('PART' in c and 'NUM' in c for c in cols):
                sheet_desc = sh; break
        except: pass
    df_desc_df = pd.read_excel(f_desc, sheet_name=sheet_desc, dtype=str)
    df_razon = pd.read_excel(ARCHIVOS_FIJOS['razon'], dtype=str)
    df_ncm = pd.read_excel(ARCHIVOS_FIJOS['ncm'], dtype=str)
    xl_proy = pd.ExcelFile(ARCHIVOS_FIJOS['proyectos'])
    df_proy = pd.read_excel(ARCHIVOS_FIJOS['proyectos'], sheet_name='Hoja2' if 'Hoja2' in xl_proy.sheet_names else xl_proy.sheet_names[0], dtype=str)

    for df in [df_madre, df_sub, df_liq, df_razon, df_proy, df_eq, df_desc_df, df_ncm]:
        df.columns = df.columns.str.strip()
    for col in df_proy.columns:
        df_proy[col] = df_proy[col].astype(str).str.replace('\xa0', ' ', regex=False).str.strip()

    ncm_validas = set(df_ncm['NCM'].str.strip().tolist())
    rs = df_razon.iloc[0]
    try: cuit_val = int(float(str(rs['CUIT'])))
    except: cuit_val = str(rs['CUIT'])
    proyectos = {str(r['CUST_CD']).strip(): r.to_dict() for _, r in df_proy.iterrows()}

    eq_col_pn = find_col(df_eq, ['PART', 'NUM']) or df_eq.columns[0]
    eq_col_eq = find_col(df_eq, ['EQUIP']) or df_eq.columns[-1]
    equipos = {norm_pn(str(r[eq_col_pn])): (str(r[eq_col_eq]).strip() if pd.notna(r[eq_col_eq]) else '') for _, r in df_eq.iterrows()}

    desc_col_pn = find_col(df_desc_df, ['PART', 'NUM']) or df_desc_df.columns[0]
    desc_col_d = find_col(df_desc_df, ['DESCRIP']) or df_desc_df.columns[-1]
    descs = {norm_pn(str(r[desc_col_pn])): (str(r[desc_col_d]).strip() if pd.notna(r[desc_col_d]) else '') for _, r in df_desc_df.iterrows()}

    derechos_liq = df_liq[df_liq['CONCEPTO'].str.contains('010', na=False)][['ITEM','PORCENTAJE']].copy()
    derechos_liq['PORCENTAJE'] = pd.to_numeric(derechos_liq['PORCENTAJE'], errors='coerce').fillna(0)
    derechos_dict = dict(zip(derechos_liq['ITEM'], derechos_liq['PORCENTAJE']))

    df_sub = df_sub.dropna(subset=['MODELO'])
    df_sub['MODELO_NORM'] = df_sub['MODELO'].astype(str).apply(norm_pn)
    df_sub['NCM10'] = df_sub['NCM'].astype(str).str[:10]
    df_sub['PCT_DERECHOS'] = df_sub['ITEM'].map(derechos_dict).fillna(0)

    col_origen = find_col(df_madre, ['ORIGIN'], ['ORIGEN'])
    sin_traduccion = []
    if col_origen:
        origenes_raw = df_madre[col_origen].dropna().astype(str).str.strip().str.upper().unique()
        sin_traduccion = sorted([o for o in origenes_raw if o and o != 'NAN' and o not in TRADUCCION_ORIGEN])

    col_inv   = find_col(df_madre, ['INVOICE', 'NUM']) or 'INVOICE_NUMBER'
    col_pn    = find_col(df_madre, ['PART', 'NUM']) or 'PART_NUMBER'
    col_qty   = find_col(df_madre, ['QTY'], ['QUANTITY'], ['CANT']) or 'QTY'
    col_cust  = find_col(df_madre, ['CUST', 'CD'], ['CUSTOMER', 'CD']) or 'CUST_CD'
    col_ext   = find_col(df_madre, ['EXTENDED', 'PRICE'], ['EXT', 'PRICE']) or 'EXTENDED_PRICE'
    col_spk   = find_col(df_madre, ['SPECIAL', 'PACK']) or 'SPECIAL PACKING'
    col_frt   = find_col(df_madre, ['FREIGHT', 'CHARGE']) or 'FREIGHT_CHARGE'
    col_bofrt = find_col(df_madre, ['BO', 'FREIGHT']) or 'BO_FREIGHT_CHARGE'
    col_emerg = find_col(df_madre, ['EMERGENCY']) or 'EMERGENCY_FILL_CHARGE_VAL'

    facturas = defaultdict(list)
    for _, row in df_madre.iterrows():
        inv = str(row.get(col_inv, '')).strip()
        if inv and inv != 'nan':
            facturas[inv].append(row.to_dict())

    preview = {}
    usados_global = set()

    for inv, items in facturas.items():
        fobs_calc = calcular_fobs(items, col_ext, col_spk, col_frt, col_bofrt, col_emerg)
        items_proc = []

        for item, fob_calc in zip(items, fobs_calc):
            pn = str(item.get(col_pn, '')).strip()
            pn_n = norm_pn(pn)
            qty_str = str(item.get(col_qty, '')).strip()
            cust_cd = str(item.get(col_cust, '')).strip()

            sub_row, item_di, estado = match_subitem(pn, qty_str, fob_calc, df_sub, usados_global)

            fob_final = ncm = valor_unit = valor_total = None
            unidad = 'UNIDAD'; ncm10 = ''; pct_der = 0.0; pn_di = pn

            if sub_row is not None:
                fob_final = round(float(sub_row['MONTO FOB']), 2)
                ncm = str(sub_row.get('NCM', '')).strip()
                ncm10 = ncm[:10]
                pct_der = float(sub_row.get('PCT_DERECHOS', 0))
                pn_di = str(sub_row.get('MODELO', pn)).strip()
                try:
                    qty_f = float(qty_str)
                    valor_unit = round(fob_final / qty_f, 2) if qty_f else fob_final
                except: valor_unit = fob_final
                valor_total = fob_final
                unidad = UNIDAD_MAP.get(str(sub_row.get('UNIDAD DECLARADA', '')).strip(), 'UNIDAD')

            ncm_ok = ncm10 in ncm_validas
            paga_der = pct_der > 0
            incluir = ncm_ok and paga_der
            motivo = '' if incluir else ('NCM no incluida en Ley Minera' if not ncm_ok else '0% derechos de importación')

            tipo_maq, marca_maq, modelo_maq = parsear_equipo(equipos.get(pn_n, ''))
            descripcion = descs.get(pn_n, descs.get(pn, ''))
            desc_ok = bool(descripcion)
            proy = proyectos.get(cust_cd, {})
            try: qty_int = int(float(qty_str)) if qty_str and qty_str != 'nan' else None
            except: qty_int = qty_str

            origen_raw = str(item.get(col_origen, '')).strip().upper() if col_origen else ''

            items_proc.append({
                'pn': pn_di, 'qty_int': qty_int, 'fob_final': fob_final, 'ncm': ncm, 'ncm10': ncm10,
                'pct_der': pct_der, 'derecho': round((fob_final or 0) * pct_der / 100, 2),
                'valor_unit': valor_unit, 'valor_total': valor_total, 'unidad': unidad,
                'tipo_maq': tipo_maq, 'marca_maq': marca_maq, 'modelo_maq': modelo_maq,
                'descripcion': descripcion, 'proy': proy,
                'origen_raw': origen_raw, 'origen': origen_raw,
                'item_di': item_di, 'estado_match': estado,
                'incluir': incluir, 'motivo': motivo, 'desc_ok': desc_ok,
            })

        validos = [i for i in items_proc if i['incluir']]
        excluidos = [i for i in items_proc if not i['incluir']]
        grupos = [validos[i:i+30] for i in range(0, max(len(validos), 1), 30)]
        grupos_info = []
        for g_idx, grupo in enumerate(grupos, 1):
            total_der = round(sum(i['derecho'] for i in grupo), 2)
            sufijo = f'.{g_idx}' if len(grupos) > 1 else ''
            grupos_info.append({'idx': g_idx, 'sufijo': sufijo, 'items': grupo, 'total_der': total_der, 'genera': len(grupo) > 0 and total_der >= 50})

        preview[inv] = {'items_proc': items_proc, 'excluidos': excluidos, 'grupos': grupos_info}

    return preview, rs, cuit_val, sin_traduccion

def aplicar_traducciones(preview, traducciones_extra):
    for inv, data in preview.items():
        for item in data['items_proc']:
            item['origen'] = traducir_pais(item['origen_raw'], traducciones_extra)
        for g in data['grupos']:
            for item in g['items']:
                item['origen'] = traducir_pais(item['origen_raw'], traducciones_extra)
        for item in data['excluidos']:
            item['origen'] = traducir_pais(item['origen_raw'], traducciones_extra)

# ── UI ────────────────────────────────────────────────────────────────────────
st.title("📋 Template CM")
st.markdown('<p class="subtitulo-app">INTERLOG Comercio Exterior — Generación automática de planillas de clasificación de mercadería</p>', unsafe_allow_html=True)
st.markdown("---")

col_main, col_config = st.columns([2.5, 1])
with col_main:
    st.markdown('<p class="seccion-titulo">📂 Archivos de la operación</p>', unsafe_allow_html=True)
    nro_ref = st.text_input("Número de referencia de la operación", placeholder="ej: 982755")
    c1, c2 = st.columns(2)
    with c1:
        f_madre = st.file_uploader(f"Excel FACAERO {nro_ref}" if nro_ref else "Excel FACAERO", type=["xlsx"])
        f_equipos = st.file_uploader(f"Excel Equipos {nro_ref}" if nro_ref else "Excel Equipos", type=["xlsx"])
    with c2:
        f_despacho = st.file_uploader(f"Excel DI {nro_ref}" if nro_ref else "Excel DI", type=["xlsx"])
        f_descripciones = st.file_uploader(f"Excel Descripciones {nro_ref}" if nro_ref else "Excel Descripciones", type=["xlsx"])

with col_config:
    st.markdown('<p class="seccion-titulo">⚙️ Configuración</p>', unsafe_allow_html=True)
    cond_merca = st.selectbox("Condición de Mercadería", ["NUEVA", "USADA"])
    st.markdown("<br>", unsafe_allow_html=True)
    archivos_din = [f_madre, f_despacho, f_equipos, f_descripciones]
    nombres_din = [f"FACAERO {nro_ref or ''}", f"DI {nro_ref or ''}", "Equipos", "Descripciones"]
    st.markdown('<p class="seccion-titulo">📋 Estado</p>', unsafe_allow_html=True)
    for f, n in zip(archivos_din, nombres_din):
        if f: st.markdown(f'<span class="badge-ok">✓ {n}</span>', unsafe_allow_html=True)
        else: st.markdown(f'<span class="badge-err">✗ {n}</span>', unsafe_allow_html=True)
    st.markdown("<br>**Archivos fijos del sistema:**")
    for nombre in ["Razón Social", "Proyectos", "NCM Ley Minera"]:
        st.markdown(f'<span class="badge-fixed">⚙ {nombre}</span>', unsafe_allow_html=True)

st.markdown("---")
todos_cargados = all(archivos_din) and bool(nro_ref)
if not nro_ref:
    st.info("📝 Ingresá el número de referencia de la operación para continuar.")

if st.button("🔍 ANALIZAR OPERACIÓN", disabled=not todos_cargados, use_container_width=True):
    with st.spinner("Analizando..."):
        try:
            preview, rs, cuit_val, sin_traduccion = procesar(f_madre, f_despacho, f_equipos, f_descripciones, cond_merca)
            st.session_state.update({
                'preview': preview, 'rs': rs, 'cuit_val': cuit_val,
                'cond_merca': cond_merca, 'nro_ref': nro_ref,
                'analizado': True, 'confirmado': False,
                'sin_traduccion': sin_traduccion,
                'traducciones_manuales': {},
            })
        except Exception as e:
            st.error(f"Error: {e}")
            import traceback; st.code(traceback.format_exc())

if st.session_state.get('analizado') and 'nro_ref' in st.session_state:
    preview = st.session_state['preview']
    nro = st.session_state['nro_ref']
    sin_traduccion = st.session_state.get('sin_traduccion', [])

    if sin_traduccion:
        st.markdown("### 🌐 Orígenes sin traducción detectados")
        st.warning(f"⚠️ Se encontraron {len(sin_traduccion)} origen(es) no reconocido(s). Completá la traducción al español antes de generar.")
        traducciones_manuales = {}
        cols_trad = st.columns(min(len(sin_traduccion), 3))
        for i, origen in enumerate(sin_traduccion):
            with cols_trad[i % 3]:
                val = st.text_input(f"🌍 {origen}", key=f"trad_{origen}", placeholder="Ej: SUIZA").strip().upper()
                traducciones_manuales[origen] = val
        st.session_state['traducciones_manuales'] = traducciones_manuales
        st.markdown("---")

    st.markdown("### 🔎 Revisión previa — Confirmá antes de generar")

    for inv, data in preview.items():
        excluidos = data['excluidos']
        grupos = data['grupos']
        total = len(data['items_proc'])
        validos = sum(len(g['items']) for g in grupos)
        st.markdown(f"**Factura {inv}** — {total} ítems totales | {validos} válidos | {len(excluidos)} excluidos")
        if excluidos:
            rows = ''.join([f"<div>⚠️ PN <b>{e['pn']}</b> | NCM {e['ncm10'] or 'sin NCM'} | {e['pct_der']}% | {e['motivo']}</div>" for e in excluidos])
            st.markdown(f'<div class="excluido-box"><div class="excluido-title">Ítems excluidos:</div>{rows}</div>', unsafe_allow_html=True)
        sin_desc = [i for g in grupos for i in g['items'] if not i.get('desc_ok')]
        if sin_desc:
            rows_desc = ''.join([f"<div>⚠️ PN <b>{i['pn']}</b> — descripción no encontrada en Excel Descripciones</div>" for i in sin_desc])
            st.markdown(f'<div class="excluido-box"><div class="excluido-title">⚠️ Descripciones faltantes:</div>{rows_desc}</div>', unsafe_allow_html=True)
        for g in grupos:
            nombre = f"TEMPLATE_{nro}_{inv}{g['sufijo']}.xlsx"
            if g['genera']:
                st.markdown(f'<span class="badge-ok">✅ {nombre} — {len(g["items"])} ítems | Derechos: USD {g["total_der"]}</span>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="no-genera-box">❌ {nombre} NO SE GENERA — {len(g["items"])} ítems | Derechos: USD {g["total_der"]} (menor a USD 50)</div>', unsafe_allow_html=True)

    st.markdown("---")
    traducciones_incompletas = [o for o, v in st.session_state.get('traducciones_manuales', {}).items() if not v]
    if traducciones_incompletas:
        st.warning(f"⚠️ Completá la traducción de: **{', '.join(traducciones_incompletas)}** para poder generar.")
    puede_confirmar = not traducciones_incompletas
    if st.button("✅ CONFIRMAR Y GENERAR TEMPLATES", use_container_width=True, disabled=not puede_confirmar):
        st.session_state['confirmado'] = True

if st.session_state.get('confirmado') and 'nro_ref' in st.session_state:
    preview = st.session_state['preview']
    rs = st.session_state['rs']
    cuit_val = st.session_state['cuit_val']
    cond_merca = st.session_state['cond_merca']
    nro = st.session_state['nro_ref']
    traducciones_extra = st.session_state.get('traducciones_manuales', {})

    aplicar_traducciones(preview, traducciones_extra)

    excel_bytes = {}
    alertas = []

    for inv, data in preview.items():
        for g in data['grupos']:
            if not g['genera']: continue
            nombre = f"{inv}{g['sufijo']}"
            filas = construir_filas(g['items'], inv, rs, cuit_val, cond_merca)
            excel_bytes[nombre] = generar_excel_bytes(filas)
            for item in g['items']:
                if item['estado_match'] == 'descarte':
                    alertas.append(f"<b>{inv}</b> | PN <code>{item['pn']}</code> → ITEM <code>{item['item_di']}</code> asignado por descarte. Verificar.")

    if alertas:
        st.markdown("### ⚠️ Alertas — Asignaciones por descarte")
        for a in alertas:
            st.markdown(f'<div class="alerta-descarte">⚠️ {a}</div>', unsafe_allow_html=True)

    st.markdown(f"### 📥 Descargar — {len(excel_bytes)} templates generados")
    cols_dl = st.columns(min(len(excel_bytes), 4))
    for i, (nombre, b) in enumerate(excel_bytes.items()):
        with cols_dl[i % 4]:
            st.download_button(label=f"📄 {nombre}", data=b,
                file_name=f"TEMPLATE_{nro}_{nombre}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{nombre}")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w') as zf:
        for nombre, b in excel_bytes.items():
            zf.writestr(f"TEMPLATE_{nro}_{nombre}.xlsx", b)
    zip_buf.seek(0)
    st.download_button(label=f"📦 DESCARGAR TODOS (ZIP) — Operación {nro}",
        data=zip_buf.getvalue(), file_name=f"TEMPLATES_{nro}.zip",
        mime="application/zip", use_container_width=True)

    todas = []
    for inv, data in preview.items():
        for g in data['grupos']:
            if g['genera']:
                todas.extend(construir_filas(g['items'], inv, rs, cuit_val, cond_merca))
            else:
                todas.extend(construir_filas_excluidos(g['items'], inv, rs, cuit_val, cond_merca))
        todas.extend(construir_filas_excluidos(data['excluidos'], inv, rs, cuit_val, cond_merca))
    if todas:
        st.download_button(label=f"📊 DESCARGAR TEMPLATE UNIFICADO — Operación {nro}",
            data=generar_excel_bytes(todas),
            file_name=f"TEMPLATE_UNIFICADO_{nro}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
