import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
import io, math, zipfile

st.set_page_config(page_title="Generador RUMP | INTERLOG", page_icon="⛏️", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
    html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
    h1 { font-family: 'IBM Plex Mono', monospace !important; color: #f0c040 !important; letter-spacing: -1px; border-bottom: 2px solid #f0c040; padding-bottom: 12px; }
    h2, h3 { font-family: 'IBM Plex Mono', monospace !important; color: #e0e0e0 !important; }
    .stButton > button { background: #f0c040 !important; color: #0f1117 !important; font-family: 'IBM Plex Mono', monospace !important; font-weight: 600 !important; border: none !important; border-radius: 2px !important; letter-spacing: 1px; }
    .stButton > button:hover { background: #ffd966 !important; }
    .badge-ok { background: #1a3a1a; color: #4caf50; padding: 4px 10px; border-radius: 2px; font-family: 'IBM Plex Mono', monospace; font-size: 12px; border: 1px solid #4caf50; display: inline-block; margin: 2px 0; }
    .badge-err { background: #3a0000; color: #f44336; padding: 4px 10px; border-radius: 2px; font-family: 'IBM Plex Mono', monospace; font-size: 12px; border: 1px solid #f44336; display: inline-block; margin: 2px 0; }
    .resumen-box { background: #1a1d27; border: 1px solid #2a2d3a; border-radius: 4px; padding: 16px; margin: 8px 0; font-family: 'IBM Plex Mono', monospace; font-size: 13px; }
</style>
""", unsafe_allow_html=True)

PAIS_MAP = {
    'USA': 'ESTADOS UNIDOS', 'UNITED STATES': 'ESTADOS UNIDOS',
    'MEXICO': 'MEXICO', 'INDIA': 'INDIA', 'CANADA': 'CANADA', 'CHINA': 'CHINA',
    'DENMARK': 'DINAMARCA', 'GERMANY': 'ALEMANIA', 'JAPAN': 'JAPON',
    'FRANCE': 'FRANCIA', 'ITALY': 'ITALIA', 'BRAZIL': 'BRASIL',
    'UK': 'REINO UNIDO', 'UNITED KINGDOM': 'REINO UNIDO', 'SWEDEN': 'SUECIA',
    'SPAIN': 'ESPAÑA', 'NETHERLANDS': 'PAISES BAJOS',
    'SOUTH KOREA': 'COREA DEL SUR', 'KOREA': 'COREA DEL SUR',
    'AUSTRALIA': 'AUSTRALIA', 'FINLAND': 'FINLANDIA',
}
UNIDAD_MAP = {'07 - UNIDAD': 'UNIDAD','01 - KILOGRAMO': 'KILOGRAMO','06 - METRO': 'METRO','10 - LITRO': 'LITRO'}

COLUMNAS = [
    'ID','InscRUMP','ActiServ','NroInsc','RazonSocial','CUIT','ImpDirecta','CondMerca',
    'AnioFabricacion','RazonSocialLeasing','CuitLeasing','SimiSira','ValorFOBTotal',
    'ProyectoMinero','Radicacion','DetalleTransitorioDeposito','ClasificacionDeArticulo',
    'TipoDeFactura','NumeroDeFactura','OrdenDeCompra','Descripcion','Cantidad',
    'UnidadMedida','PosicionArancelaria','Marca','Modelo','NroDeSerie','ValorUnitario',
    'ValorTotalItem','SerieDelMotor','MarcaDelMotor','ModeloMotor','TipoPlantaDestino',
    'FinalidadDeUso','CodigoParte','TipoDeMaquina','MarcaMaquina','ModeloMaquina',
    'Expedientes Escalonados','Proveedor','PaisOrigenDeLaMercaderia','Observaciones',
    'ITEM_DESPACHO'
]
SIN_COLOR = {
    'ID','InscRUMP','ActiServ','NroInsc','RazonSocial','CUIT','ImpDirecta','CondMerca',
    'SimiSira','ProyectoMinero','Radicacion','ClasificacionDeArticulo','TipoDeFactura',
    'Observaciones','ITEM_DESPACHO'
}

def safe_float(v):
    try:
        f = float(v); return 0.0 if math.isnan(f) else f
    except: return 0.0

def traducir_pais(p): return PAIS_MAP.get(str(p).strip().upper(), str(p).strip().upper())

def parsear_equipo(eq):
    if not eq or eq.strip() in ('', 'nan'): return '', 'CAT', ''
    eq = eq.strip()
    if ' - ' in eq:
        p = eq.split(' - ', 1); return p[0].strip(), 'CAT', p[1].strip()
    tokens = eq.split()
    for i, t in enumerate(tokens):
        if t.upper() in ['CAT', 'CATERPILLAR']:
            return ' '.join(tokens[:i]).strip(), 'CAT', ' '.join(tokens[i+1:]).strip()
    return (' '.join(tokens[:-1]).strip(), 'CAT', tokens[-1].strip()) if len(tokens) > 1 else (eq, 'CAT', '')

def calcular_fobs(items):
    total_ext = sum(safe_float(i.get('EXTENDED_PRICE')) for i in items)
    total_gastos = sum(
        safe_float(i.get('SPECIAL PACKING')) + safe_float(i.get('FREIGHT_CHARGE')) +
        safe_float(i.get('BO_FREIGHT_CHARGE')) + safe_float(i.get('EMERGENCY_FILL_CHARGE_VAL'))
        for i in items
    )
    return [round(safe_float(i.get('EXTENDED_PRICE')) + (safe_float(i.get('EXTENDED_PRICE')) * total_gastos / total_ext if total_ext else 0), 2) for i in items]

def match_fob(pn, qty, fob_calc, subitem_df, usados):
    pn_str = str(pn).strip()
    try: qty_val = float(qty)
    except: return None, None, 'sin_match'
    mask = (subitem_df['MODELO'].astype(str).str.strip() == pn_str) & (subitem_df['CANTIDAD'].astype(float) == qty_val)
    disponibles = subitem_df[mask & ~subitem_df.index.isin(usados)].copy()
    if disponibles.empty: return None, None, 'sin_match'
    match = disponibles[(disponibles['MONTO FOB'].astype(float) - fob_calc).abs() <= 0.02]
    if not match.empty:
        idx = match.index[0]; usados.add(idx); return match.loc[idx], str(match.loc[idx]['ITEM']), 'unico'
    idx = disponibles.index[0]; usados.add(idx); return disponibles.loc[idx], str(disponibles.loc[idx]['ITEM']), 'descarte'

def generar_excel_bytes(filas):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'general'
    FILL_AM = PatternFill('solid', fgColor='FFFF00')
    FH = Font(name='Arial', size=11, bold=True)
    FD = Font(name='Calibri', size=11)
    for col_idx, col_name in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = FH; cell.alignment = Alignment(horizontal='left', vertical='center')
        if col_name not in SIN_COLOR: cell.fill = FILL_AM
    for row_idx, fila in enumerate(filas, 2):
        for col_idx, col_name in enumerate(COLUMNAS, 1):
            val = fila.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FD; cell.alignment = Alignment(vertical='center')
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def procesar(f_madre, f_despacho, f_razon, f_proyectos, f_equipos, f_desc, f_ncm, cond_merca):
    df_madre = pd.read_excel(f_madre, sheet_name='Hoja2', dtype=str)
    df_subitem = pd.read_excel(f_despacho, sheet_name='Subitem', dtype=str)
    df_razon = pd.read_excel(f_razon, dtype=str)
    df_ncm_raw = pd.read_excel(f_ncm, dtype=str)

    # Proyectos — detectar solapa
    xl_proy = pd.ExcelFile(f_proyectos)
    sheet_proy = 'Hoja2' if 'Hoja2' in xl_proy.sheet_names else xl_proy.sheet_names[0]
    df_proy = pd.read_excel(f_proyectos, sheet_name=sheet_proy, dtype=str)

    # Equipos — detectar solapa
    xl_eq = pd.ExcelFile(f_equipos)
    sheet_eq = 'data' if 'data' in xl_eq.sheet_names else xl_eq.sheet_names[0]
    df_eq = pd.read_excel(f_equipos, sheet_name=sheet_eq, dtype=str)

    # Descripciones — detectar solapa
    xl_desc = pd.ExcelFile(f_desc)
    sheet_desc = 'Hoja2' if 'Hoja2' in xl_desc.sheet_names else xl_desc.sheet_names[0]
    df_desc_df = pd.read_excel(f_desc, sheet_name=sheet_desc, dtype=str)

    for df in [df_madre, df_subitem, df_razon, df_proy, df_eq, df_desc_df, df_ncm_raw]:
        df.columns = df.columns.str.strip()

    for col in df_proy.columns:
        df_proy[col] = df_proy[col].astype(str).str.replace('\xa0', ' ', regex=False).str.strip()

    ncm_validas = set(df_ncm_raw['NCM'].str.strip().tolist())
    rs = df_razon.iloc[0]
    proyectos = {str(r['CUST_CD']).strip(): r.to_dict() for _, r in df_proy.iterrows()}

    eq_col_pn = 'Part number' if 'Part number' in df_eq.columns else df_eq.columns[0]
    eq_col_eq = 'Equipos' if 'Equipos' in df_eq.columns else df_eq.columns[-1]
    equipos = {str(r[eq_col_pn]).strip(): (str(r[eq_col_eq]).strip() if pd.notna(r[eq_col_eq]) else '') for _, r in df_eq.iterrows()}

    desc_col_pn = 'PART_NUMBER' if 'PART_NUMBER' in df_desc_df.columns else df_desc_df.columns[0]
    desc_col_d = 'DESCRIPCION' if 'DESCRIPCION' in df_desc_df.columns else df_desc_df.columns[-1]
    descs = {str(r[desc_col_pn]).strip(): (str(r[desc_col_d]).strip() if pd.notna(r[desc_col_d]) else '') for _, r in df_desc_df.iterrows()}

    try: cuit_val = int(float(str(rs['CUIT'])))
    except: cuit_val = str(rs['CUIT'])

    facturas = defaultdict(list)
    for _, row in df_madre.iterrows():
        inv = str(row.get('INVOICE_NUMBER', '')).strip()
        if inv and inv != 'nan':
            facturas[inv].append(row.to_dict())

    resultados = {}
    alertas = []
    usados_global = set()

    for inv, items in facturas.items():
        fobs_calc = calcular_fobs(items)
        filas_candidatas = []

        for item, fob_calc in zip(items, fobs_calc):
            pn = str(item.get('PART_NUMBER', '')).strip()
            qty_str = str(item.get('QTY', '')).strip()
            cust_cd = str(item.get('CUST_CD', '')).strip()

            sub_row, item_despacho, estado = match_fob(pn, qty_str, fob_calc, df_subitem, usados_global)

            fob_final = ncm = valor_unit = valor_total = None
            unidad = 'UNIDAD'; ncm_10 = ''

            if sub_row is not None:
                fob_final = round(float(sub_row['MONTO FOB']), 2)
                ncm = str(sub_row.get('NCM', '')).strip()
                ncm_10 = ncm[:10]
                try:
                    qty_f = float(qty_str)
                    valor_unit = round(fob_final / qty_f, 2) if qty_f else fob_final
                except: valor_unit = fob_final
                valor_total = fob_final
                unidad = UNIDAD_MAP.get(str(sub_row.get('UNIDAD DECLARADA', '')).strip(), 'UNIDAD')
                if estado == 'descarte':
                    alertas.append({'factura': inv, 'pn': pn, 'fob_calc': fob_calc,
                                    'fob_despacho': fob_final, 'diff': round(fob_final - fob_calc, 2),
                                    'item_despacho': item_despacho})

            if ncm_10 not in ncm_validas:
                continue

            tipo_maq, marca_maq, modelo_maq = parsear_equipo(equipos.get(pn, ''))
            descripcion = descs.get(pn, str(item.get('PART_NAME', '')).strip())
            proy = proyectos.get(cust_cd, {})

            try: qty_int = int(float(qty_str)) if qty_str and qty_str != 'nan' else None
            except: qty_int = qty_str

            filas_candidatas.append({
                'pn': pn, 'qty_int': qty_int, 'fob_final': fob_final, 'ncm': ncm,
                'valor_unit': valor_unit, 'valor_total': valor_total, 'unidad': unidad,
                'tipo_maq': tipo_maq, 'marca_maq': marca_maq, 'modelo_maq': modelo_maq,
                'descripcion': descripcion, 'proy': proy,
                'origen': traducir_pais(str(item.get('PART_ORIGIN', '')).strip()),
                'item_despacho': item_despacho,
            })

        if not filas_candidatas:
            continue

        valor_fob_total = round(sum(f['fob_final'] for f in filas_candidatas if f['fob_final']), 2)
        filas = []
        for f in filas_candidatas:
            proy = f['proy']
            filas.append({
                'ID': 0,
                'InscRUMP': str(rs['InscRUMP']), 'ActiServ': str(rs['ActiServ']),
                'NroInsc': str(rs['NroInsc']), 'RazonSocial': str(rs['RazonSocial']),
                'CUIT': cuit_val, 'ImpDirecta': str(rs['ImpDirecta']),
                'CondMerca': cond_merca,
                'AnioFabricacion': None, 'RazonSocialLeasing': None, 'CuitLeasing': None, 'SimiSira': None,
                'ValorFOBTotal': valor_fob_total,
                'ProyectoMinero': str(proy.get('ProyectoMinero', '')).upper() if proy else '',
                'Radicacion': str(proy.get('Radicacion', '')).upper() if proy else '',
                'DetalleTransitorioDeposito': str(proy.get('DetalleTransitorioDeposito', '')) if proy else '',
                'ClasificacionDeArticulo': str(rs['ClasificacionDeArticulo']),
                'TipoDeFactura': 'DEFINITIVA', 'NumeroDeFactura': inv, 'OrdenDeCompra': None,
                'Descripcion': f['descripcion'], 'Cantidad': f['qty_int'], 'UnidadMedida': f['unidad'],
                'PosicionArancelaria': f['ncm'], 'Marca': 'CATERPILLAR', 'Modelo': 'NO POSEE',
                'NroDeSerie': 'NO POSEE', 'ValorUnitario': f['valor_unit'], 'ValorTotalItem': f['valor_total'],
                'SerieDelMotor': None, 'MarcaDelMotor': None, 'ModeloMotor': None,
                'TipoPlantaDestino': None, 'FinalidadDeUso': None,
                'CodigoParte': f['pn'], 'TipoDeMaquina': f['tipo_maq'],
                'MarcaMaquina': f['marca_maq'], 'ModeloMaquina': f['modelo_maq'],
                'Expedientes Escalonados': 'no', 'Proveedor': 'CATERPILLAR',
                'PaisOrigenDeLaMercaderia': f['origen'], 'Observaciones': None,
                'ITEM_DESPACHO': f['item_despacho'],
            })
        resultados[inv] = filas

    return resultados, alertas


# ── UI ────────────────────────────────────────────────────────────────────────
st.title("⛏️ GENERADOR RUMP")
st.markdown("**INTERLOG Comercio Exterior** — Sistema de generación de planillas RUMP")
st.markdown("---")

col1, col2 = st.columns([2, 1])

with col1:
    st.markdown("### 📂 Archivos de entrada")
    c1, c2 = st.columns(2)
    with c1:
        f_madre = st.file_uploader("Excel Madre (FANU...)", type=["xlsx"])
        f_despacho = st.file_uploader("Excel Despacho (26001IC...)", type=["xlsx"])
        f_razon = st.file_uploader("Excel Razón Social", type=["xlsx"])
        f_ncm = st.file_uploader("Excel NCM Ley Minera", type=["xlsx"])
    with c2:
        f_proyectos = st.file_uploader("Excel Proyectos Mineros", type=["xlsx"])
        f_equipos = st.file_uploader("Excel Equipos", type=["xlsx"])
        f_descripciones = st.file_uploader("Excel Descripciones", type=["xlsx"])

with col2:
    st.markdown("### ⚙️ Configuración")
    cond_merca = st.selectbox("Condición de Mercadería", ["NUEVA", "USADA"])
    st.markdown("<br>", unsafe_allow_html=True)
    archivos = [f_madre, f_despacho, f_razon, f_proyectos, f_equipos, f_descripciones, f_ncm]
    nombres = ["Excel Madre", "Despacho", "Razón Social", "Proyectos", "Equipos", "Descripciones", "NCM"]
    st.markdown("**Estado de archivos:**")
    for f, n in zip(archivos, nombres):
        if f:
            st.markdown(f'<span class="badge-ok">✓ {n}</span>', unsafe_allow_html=True)
        else:
            st.markdown(f'<span class="badge-err">✗ {n}</span>', unsafe_allow_html=True)

st.markdown("---")

if st.button("🚀 GENERAR PLANILLAS", disabled=not all(archivos), use_container_width=True):
    with st.spinner("Procesando..."):
        try:
            resultados, alertas = procesar(f_madre, f_despacho, f_razon, f_proyectos, f_equipos, f_descripciones, f_ncm, cond_merca)
            st.session_state['resultados'] = resultados
            st.session_state['alertas'] = alertas
            st.session_state['procesado'] = True
        except Exception as e:
            st.error(f"Error: {e}")
            import traceback; st.code(traceback.format_exc())

if st.session_state.get('procesado'):
    resultados = st.session_state['resultados']
    alertas = st.session_state['alertas']

    if alertas:
        st.markdown("### ⚠️ Alertas — Asignaciones por Descarte")
        for a in alertas:
            st.warning(f"**{a['factura']}** | PN `{a['pn']}` → ITEM `{a['item_despacho']}` por descarte | FOB calc: {a['fob_calc']} | FOB despacho: {a['fob_despacho']} | diff: **{a['diff']}**")

    st.markdown("### 📊 Resumen")
    total_items = sum(len(v) for v in resultados.values())
    cols = st.columns(min(len(resultados), 4))
    for i, (inv, filas) in enumerate(resultados.items()):
        with cols[i % 4]:
            st.markdown(f'<div class="resumen-box"><b>{inv}</b><br>{len(filas)} ítems<br>FOB total: {filas[0]["ValorFOBTotal"] if filas else "-"}</div>', unsafe_allow_html=True)
    st.markdown(f"**{len(resultados)} facturas | {total_items} ítems totales**")
    st.markdown("---")

    st.markdown("### 📥 Descargar")
    excel_bytes = {}
    cols_dl = st.columns(min(len(resultados), 4))
    for i, (inv, filas) in enumerate(resultados.items()):
        b = generar_excel_bytes(filas)
        excel_bytes[inv] = b
        with cols_dl[i % 4]:
            st.download_button(label=f"📄 {inv}", data=b, file_name=f"{inv}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{inv}")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w') as zf:
        for inv, b in excel_bytes.items():
            zf.writestr(f"{inv}.xlsx", b)
    zip_buf.seek(0)
    st.download_button(label="📦 DESCARGAR TODOS (ZIP)", data=zip_buf.getvalue(),
                       file_name="RUMP_planillas.zip", mime="application/zip", use_container_width=True)
